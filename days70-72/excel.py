from openpyxl import load_workbook


def main():
    wb = load_workbook('Financial Sample.xlsx')
    print('Sheets in the Workbook Financial Sample.xlsx')
    print(wb.sheetnames)

    ws1 = wb['Finances 2017']
    print('Value of Cell B3 in Finances 2017 sheet')
    print(ws1['B3'].value)

    print('Total profit: Rows 2 to 6')
    total_profit = 0
    for col in list('L'):
        for row in range(2, 6):
            cell = f'{col}{row}'
            total_profit += float(ws1[cell].value)
    print(total_profit)

    print('Max row')
    maxrow = ws1.max_row
    print(maxrow)

    print('Printing the Countries.')
    for row in range(2, maxrow):
        cell = f'B{row}'
        print(ws1[cell].value)

    print('Get the total of profits and save back to the worksheet')
    cell = f'L{(maxrow + 2)}'
    ws1[cell] = f'=SUM(L2:L{maxrow - 1})'
    wb.save('Financial Sample.xlsx')


if __name__ == '__main__':
    main()
